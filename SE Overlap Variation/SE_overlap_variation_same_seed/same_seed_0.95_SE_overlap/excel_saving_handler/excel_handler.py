# excel_handler.py
import openpyxl
from openpyxl.styles import Font, Border, Side
import os
import pandas as pd

def get_workbook_and_sheet(file_path, sheet_name, headers=None):
    """
    Retrieves an openpyxl workbook and a specific sheet.
    If the file or sheet does not exist, they will be created.
    Headers are appended and styled only if a new sheet is created.
    """
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()
        # If a new workbook is created, and it still has the default 'Sheet', remove it
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
        if headers is not None: # Check if headers are provided
            sheet.append(headers)
            apply_header_style(sheet)
            
    return workbook, sheet

def apply_header_style(sheet):
    """Applies bold font and thin borders to the first row (headers) of a sheet."""
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    if sheet.max_row > 0: # Check if there's at least one row (headers)
        for cell in sheet[1]: # sheet[1] refers to the first row
            cell.font = bold_font
            cell.border = thin_border

def save_data_to_excel(workbook, sheet, data_rows, file_path):
    """
    Appends multiple rows of data to the specified sheet and saves the workbook.
    Converts tuples to strings before saving.
    
    Args:
        workbook: The openpyxl workbook object.
        sheet: The openpyxl sheet object to append data to.
        data_rows: A list of lists, where each inner list is a row of data.
        file_path: The path where the workbook should be saved.
    """
    processed_rows = []
    for row_data in data_rows:
        processed_row = []
        for item in row_data:
            if isinstance(item, tuple):
                processed_row.append(str(item)) # Convert tuple to string
            elif isinstance(item, list):
                # Convert list to string as well, to avoid similar issues
                processed_row.append(str(item)) 
            else:
                processed_row.append(item)
        processed_rows.append(processed_row)

    for row_data in processed_rows:
        sheet.append(row_data)
    workbook.save(file_path)

def copy_sheet_to_final_excel(temp_excel_path, sheet_name_to_copy, final_writer):
    """
    Copies a specific sheet from a temporary Excel file to the final Excel writer.
    
    Args:
        temp_excel_path (str): Path to the temporary Excel file.
        sheet_name_to_copy (str): Name of the sheet to copy.
        final_writer (pd.ExcelWriter): The pandas ExcelWriter object for the final file.
    """
    try:
        if os.path.exists(temp_excel_path):
            temp_workbook = openpyxl.load_workbook(temp_excel_path)
            if sheet_name_to_copy in temp_workbook.sheetnames:
                temp_sheet = temp_workbook[sheet_name_to_copy]
                
                # Create a new DataFrame from the sheet data, starting from the second row (after headers)
                # and using the first row as columns
                data = temp_sheet.iter_rows(min_row=2, values_only=True)
                columns = [cell.value for cell in temp_sheet[1]]
                
                df_to_copy = pd.DataFrame(data, columns=columns)
                df_to_copy.to_excel(final_writer, sheet_name=sheet_name_to_copy, index=False)
            else:
                print(f"Warning: Sheet '{sheet_name_to_copy}' not found in '{temp_excel_path}'.")
        else:
            print(f"Error: Temporary file not found: '{temp_excel_path}'. Cannot copy sheet '{sheet_name_to_copy}'.")
    except Exception as e:
        print(f"Error copying sheet '{sheet_name_to_copy}' from '{temp_excel_path}' to final Excel: {e}")